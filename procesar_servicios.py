#Angel Saldaña Aguilar
#Patricio Emiliano Garay Elizondo
import subprocess
import openpyxl
from openpyxl import Workbook
import os

def ejecutar_powershell(script_ps1):
    try:
        proceso = subprocess.run(["powershell", "-ExecutionPolicy", "Bypass", "-File", script_ps1], 
                                 capture_output=True, text=True, check=True)
        salida = proceso.stdout
        return salida
    except subprocess.CalledProcessError as e:
        print(f"Error al ejecutar el script {script_ps1}: {e}")
        print(f"Salida del error: {e.output}")
        return None

def guardar_en_excel(salida, archivo_excel):
    try:
        if os.path.exists(archivo_excel):
            wb = openpyxl.load_workbook(archivo_excel)
        else:
            wb = Workbook()
        # Selecciona la hoja activa
        ws = wb.active
        # Agrega encabezado si el archivo es nuevo
        if ws.max_row == 1:
            ws.append(["Salida PowerShell"])
        # Divide la salida del script en líneas y escribe cada una en una fila
        for linea in salida.splitlines():
            ws.append([linea])
        # Guardar el archivo Excel
        wb.save(archivo_excel)
        print(f"Datos guardados en {archivo_excel} con exito.")
    except Exception as e:
        print(f"Error al guardar en Excel: {e}")
def main():
    script_ps1 = "monitor_servicios.ps1"
    archivo_excel = "salida_powershell.xlsx"
    if not os.path.exists(script_ps1):
        print(f"El archivo {script_ps1} no existe en el directorio actual.")
        return
    salida = ejecutar_powershell(script_ps1)
    if salida:
        guardar_en_excel(salida, archivo_excel)
if __name__ == "__main__":
    main()